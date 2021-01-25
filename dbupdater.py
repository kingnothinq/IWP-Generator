from openpyxl import load_workbook
from pathlib import Path
from re import compile
from tinydb import TinyDB


def get_slices(sheet):
    """Find devices in XLS."""

    row_start = 1
    while True:
        family_cell = sheet.cell(row=row_start, column=2)
        if family_cell.value == 'Quanta 5':
            row_old_start = row_start
            row_start = row_start + 24
            row_end = row_start - 2
        elif family_cell.value == 'Quanta 6':
            row_old_start = row_start
            row_start = row_start + 24
            row_end = row_start - 2
        elif family_cell.value == 'Quanta 70':
            row_old_start = row_start
            row_start = row_start + 9
            row_end = row_start - 2
        elif family_cell.value == 'InfiLINK XG 1000':
            row_old_start = row_start
            row_start = row_start + 15
            row_end = row_start - 2
        elif family_cell.value == 'InfiLINK XG 500':
            row_old_start = row_start
            row_start = row_start + 15
            row_end = row_start - 2
        elif family_cell.value == 'InfiLINK 2x2 PRO':
            row_old_start = row_start
            row_start = row_start + 18
            row_end = row_start - 2
        elif family_cell.value == 'InfiLINK 2x2 LITE':
            row_old_start = row_start
            row_start = row_start + 18
            row_end = row_start - 2
        elif family_cell.value == 'InfiLINK Evolution':
            row_old_start = row_start
            row_start = row_start + 21
            row_end = row_start - 2
        elif family_cell.value == 'Axion 28':
            row_old_start = row_start
            row_start = row_start + 9
            row_end = row_start - 2
        elif family_cell.value is None:
            break
        yield row_old_start, row_end


def analyze_slice(start, end, sheet):
    """Parse data from XLS."""

    capacity = {}
    availability = {'99.90': {}, '99.99': {}}
    for row in sheet.iter_rows(min_row=start, min_col=1, max_row=end, max_col=17):
        row_text = list(filter(lambda x: x is not None, [cell.value for cell in row]))
        if row_text[0] == 'Family':
            family = row_text[1]
        elif row_text[0] == 'Device':
            name = row_text[1]
        elif row_text[0] == 'Capacity, Mbps':
            capacity[f'{row_text[1]}'] = {f'MCS{id}': int(item) for id, item in enumerate(row_text[2:])}
        elif row_text[0] == 'Availability, 99.90%':
            availability['99.90'][row_text[1]] = {f'MCS{id}': float(item) for id, item in enumerate(row_text[2:])}
        elif row_text[0] == 'Availability, 99.99%':
            availability['99.99'][row_text[1]] = {f'MCS{id}': float(item) for id, item in enumerate(row_text[2:])}
    if ('-E' in name) or ('Um' in name) or ('Omx' in name) or ('Lmn' in name) or ('STE' in name):
        type = 'external'
        model = name.split(' + ')[0]
        antenna = name.split(' + ')[1]
        cable = 'CAB-RF-1M'
    else:
        type = 'internal'
        model = name
        antenna = None
        cable = None

    return family, name, model, type, antenna, cable, capacity, availability


def write_db(table, family, name, model, type, antenna, cable, capacity, availability):
    """Insert data to the database."""

    result = {'Family': family,
              'Name': name,
              'Model': model,
              'Type': type,
              'Antenna': antenna,
              'RF Cable': cable,
              'Capacity': capacity,
              'Availability': availability
              }
    table.insert(result)


def update_database(file_db, file_xls):
    """Write data to the database. From XLSX to JSON (*.DB)."""

    db = TinyDB(file_db)
    db.drop_tables()
    wb = load_workbook(filename=file_xls)
    for sheet in wb:
        table = db.table(sheet.title)
        for slice in get_slices(sheet):
            device = analyze_slice(*slice, sheet)
            write_db(table, *device)


if __name__ == "__main__":
    file_db = Path.cwd() / 'devices.db'
    file_xls = Path.cwd() / 'devices.xlsx'
    update_database(file_db, file_xls)
